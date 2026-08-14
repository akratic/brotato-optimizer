import csv
import re
import shutil

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

XLSX_PATH = "Copy of ArosRising's Brotato MultiTool 1.4.xlsx"
CSV_PATH = "weapons.csv"
SHEET = "WPNStats"

SCALING_STAT_COLUMNS = {
    "Prim Stat": 35,   # AI
    "2 Scaling Stat": 36,  # AJ
    "3 Scaling Stat": 37,  # AK
}


def build_name_by_cell(wb):
    name_by_cell = {}
    for name, defn in wb.defined_names.items():
        if name.startswith("DPS_") and "!" in defn.attr_text:
            sheet, coord = defn.attr_text.split("!")
            name_by_cell[(sheet.strip("'"), coord.replace("$", ""))] = name
    return name_by_cell


def parse_scaling_formula(raw, name_by_cell):
    """Resolve a WPNStats Prim/2nd/3rd-Scaling-Stat cell to (named_stat, multiplier)."""
    if isinstance(raw, ArrayFormula):
        raw = raw.text
    if raw is None or isinstance(raw, (int, float)):
        return None, 1.0  # literal constant (e.g. 0) => no scaling stat
    if not isinstance(raw, str) or not raw.startswith("="):
        return None, 1.0
    expr = raw[1:]

    m = re.match(r"^(DPS_[A-Za-z_]+)(?:\*(\d+(?:\.\d+)?))?$", expr)
    if m:
        return m.group(1), float(m.group(2)) if m.group(2) else 1.0

    m = re.match(r"^'DPS Calculator'!\$?([A-Z]+)\$?(\d+)(?:\*(\d+(?:\.\d+)?))?$", expr)
    if m:
        col, row, mult = m.group(1), m.group(2), m.group(3)
        name = name_by_cell[("DPS Calculator", f"{col}{row}")]
        return name, float(mult) if mult else 1.0

    raise ValueError(f"Unrecognized scaling-stat formula: {raw!r}")


def clean_value(v):
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def main():
    shutil.copy(CSV_PATH, "weapons_original_export.csv.bak")

    wb_formulas = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    wb_values = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    name_by_cell = build_name_by_cell(wb_formulas)

    ws_formulas = wb_formulas[SHEET]
    ws_values = wb_values[SHEET]

    header = [ws_formulas.cell(row=1, column=c).value for c in range(1, ws_formulas.max_column + 1)]

    out_header = []
    for col_idx, name in enumerate(header, start=1):
        out_header.append(name)
        if name in SCALING_STAT_COLUMNS:
            out_header.append(f"{name} Multiplier")

    rows_out = [out_header]
    for r in range(2, ws_formulas.max_row + 1):
        weapon = ws_values.cell(row=r, column=1).value
        if weapon is None:
            continue
        out_row = []
        for col_idx, name in enumerate(header, start=1):
            if name in SCALING_STAT_COLUMNS:
                raw_formula = ws_formulas.cell(row=r, column=col_idx).value
                stat_name, mult = parse_scaling_formula(raw_formula, name_by_cell)
                out_row.append(stat_name if stat_name else "")
                out_row.append(mult)
            else:
                out_row.append(clean_value(ws_values.cell(row=r, column=col_idx).value))
        rows_out.append(out_row)

    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out) - 1} weapon rows to {CSV_PATH}")
    print(f"Original export backed up to weapons_original_export.csv.bak")


if __name__ == "__main__":
    main()
